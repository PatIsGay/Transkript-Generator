"""
Vimeo Transkript Generator
===========================
Liest Vimeo-Links aus einer Excel-Datei, laedt die Audio-Spuren herunter,
transkribiert sie mit faster-whisper und speichert die Ergebnisse als Excel.

Nutzung:
    python transcribe.py                  # Alles ausfuehren
    python transcribe.py --skip-download  # Nur Transkription (Audio muss vorhanden sein)
    python transcribe.py --skip-transcribe # Nur Audio-Download
    python transcribe.py --model large-v3  # Anderes Whisper-Modell verwenden
"""

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import openpyxl

# ============================================================================
# KONFIGURATION - hier bei Bedarf anpassen
# ============================================================================

EXCEL_FILENAME = "Übungen_Reihenfolge_Modul5_6_v5.xlsx"
SHEET_NAME = "Master"
OUTPUT_DIR = "output"
AUDIO_DIR = os.path.join(OUTPUT_DIR, "audio")
PROGRESS_FILE = os.path.join(OUTPUT_DIR, "progress.json")
OUTPUT_EXCEL = os.path.join(OUTPUT_DIR, "ergebnisse.xlsx")
OUTPUT_CSV = os.path.join(OUTPUT_DIR, "ergebnisse.csv")

DEFAULT_MODEL = "small"  # "tiny", "small", "medium", "large-v3"
LANGUAGE = "de"

# Excel-Spalten (1-basiert)
COL_ORDER = 1
COL_MODUL = 2
COL_BEREICH = 3
COL_KATEGORIE = 4
COL_UEBUNG = 5
COL_UEBUNGSTYP = 6
COL_KPI = 7
COL_VIDEO_KURZ = 8
COL_VIDEO_LANG = 9

# ============================================================================
# HILFSFUNKTIONEN
# ============================================================================


def extract_vimeo_id(url: str) -> str | None:
    """Extrahiert die numerische Vimeo-ID aus einer URL."""
    if not url:
        return None
    m = re.search(r"vimeo\.com/(\d+)", str(url))
    return m.group(1) if m else None


def clean_vimeo_url(url: str) -> str:
    """Entfernt ?share=copy und normalisiert die URL."""
    if not url:
        return ""
    return str(url).split("?")[0].strip()


def load_progress(progress_file: str) -> dict:
    """Laedt den Fortschritt aus der JSON-Datei."""
    if os.path.exists(progress_file):
        with open(progress_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"downloaded": {}, "transcribed": {}}


def save_progress(progress_file: str, progress: dict):
    """Speichert den Fortschritt in die JSON-Datei."""
    with open(progress_file, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def format_duration(seconds: float) -> str:
    """Formatiert Sekunden als mm:ss."""
    m, s = divmod(int(seconds), 60)
    h, m = divmod(m, 60)
    if h > 0:
        return f"{h}h {m:02d}m {s:02d}s"
    return f"{m}m {s:02d}s"


# ============================================================================
# PHASE 1: EXCEL EINLESEN
# ============================================================================


def read_excel(excel_path: str) -> list[dict]:
    """
    Liest die Excel-Datei und gibt eine flache Liste von Eintraegen zurueck.
    Jeder Vimeo-Link wird zu einem eigenen Eintrag (kurz und lang getrennt).
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[SHEET_NAME]

    entries = []
    for row in range(2, ws.max_row + 1):
        order = ws.cell(row=row, column=COL_ORDER).value
        modul = ws.cell(row=row, column=COL_MODUL).value
        bereich = ws.cell(row=row, column=COL_BEREICH).value
        kategorie = ws.cell(row=row, column=COL_KATEGORIE).value
        uebung = ws.cell(row=row, column=COL_UEBUNG).value

        video_kurz = ws.cell(row=row, column=COL_VIDEO_KURZ).value
        video_lang = ws.cell(row=row, column=COL_VIDEO_LANG).value

        base = {
            "row": row,
            "order": order,
            "modul": str(modul) if modul else "",
            "bereich": str(bereich) if bereich else "",
            "kategorie": str(kategorie) if kategorie else "",
            "uebung": str(uebung) if uebung else "",
        }

        if video_kurz and "vimeo" in str(video_kurz).lower():
            url = clean_vimeo_url(video_kurz)
            vimeo_id = extract_vimeo_id(url)
            entries.append({
                **base,
                "link_typ": "kurz",
                "url": str(video_kurz).strip(),
                "url_clean": url,
                "vimeo_id": vimeo_id,
            })

        if video_lang and "vimeo" in str(video_lang).lower():
            url = clean_vimeo_url(video_lang)
            vimeo_id = extract_vimeo_id(url)
            entries.append({
                **base,
                "link_typ": "lang",
                "url": str(video_lang).strip(),
                "url_clean": url,
                "vimeo_id": vimeo_id,
            })

    wb.close()
    return entries


# ============================================================================
# PHASE 2: AUDIO-DOWNLOAD
# ============================================================================


def download_audio(entries: list[dict], progress: dict) -> dict:
    """
    Laedt fuer jede einzigartige Vimeo-ID die Audio-Spur herunter.
    Gibt ein Mapping {vimeo_id: audio_filepath_or_error} zurueck.
    """
    import yt_dlp

    unique_ids = {}
    for e in entries:
        vid = e["vimeo_id"]
        if vid and vid not in unique_ids:
            unique_ids[vid] = e["url_clean"]

    total = len(unique_ids)
    print(f"\n{'='*60}")
    print(f"PHASE 2: Audio-Download ({total} einzigartige Videos)")
    print(f"{'='*60}\n")

    downloaded = progress.get("downloaded", {})
    skipped = 0
    failed = 0
    new_downloads = 0

    for i, (vimeo_id, url) in enumerate(unique_ids.items(), 1):
        audio_path = os.path.join(AUDIO_DIR, f"{vimeo_id}.mp3")

        if vimeo_id in downloaded and downloaded[vimeo_id].get("status") == "ok":
            if os.path.exists(downloaded[vimeo_id]["path"]):
                skipped += 1
                print(f"  [{i}/{total}] {vimeo_id} -- bereits vorhanden, uebersprungen")
                continue

        print(f"  [{i}/{total}] Lade herunter: {vimeo_id} ...", end=" ", flush=True)

        ydl_opts = {
            "format": "bestaudio/best",
            "outtmpl": os.path.join(AUDIO_DIR, f"{vimeo_id}.%(ext)s"),
            "postprocessors": [{
                "key": "FFmpegExtractAudio",
                "preferredcodec": "mp3",
                "preferredquality": "128",
            }],
            "quiet": True,
            "no_warnings": True,
            "noprogress": True,
        }

        try:
            with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                ydl.download([url])

            if os.path.exists(audio_path):
                size_mb = os.path.getsize(audio_path) / (1024 * 1024)
                downloaded[vimeo_id] = {"status": "ok", "path": audio_path}
                new_downloads += 1
                print(f"OK ({size_mb:.1f} MB)")
            else:
                downloaded[vimeo_id] = {"status": "error", "error": "Datei nicht gefunden nach Download"}
                failed += 1
                print("FEHLER (Datei nicht erstellt)")

        except Exception as ex:
            downloaded[vimeo_id] = {"status": "error", "error": str(ex)[:200]}
            failed += 1
            print(f"FEHLER: {str(ex)[:80]}")

        progress["downloaded"] = downloaded
        save_progress(PROGRESS_FILE, progress)

    print(f"\n  Ergebnis: {new_downloads} neu | {skipped} uebersprungen | {failed} fehlgeschlagen")
    return downloaded


# ============================================================================
# PHASE 3: TRANSKRIPTION
# ============================================================================


def detect_device() -> tuple[str, int]:
    """Erkennt ob CUDA-GPU verfuegbar ist. Gibt (device, compute_type_hint) zurueck."""
    try:
        import torch
        if torch.cuda.is_available():
            gpu_name = torch.cuda.get_device_name(0)
            print(f"  GPU erkannt: {gpu_name}")
            return "cuda", 0
    except ImportError:
        pass

    try:
        from faster_whisper.utils import get_assets_path
        import ctranslate2
        if "cuda" in ctranslate2.get_supported_compute_types("cuda"):
            print("  CUDA via CTranslate2 erkannt")
            return "cuda", 0
    except Exception:
        pass

    print("  Keine GPU erkannt -- verwende CPU")
    return "cpu", 0


def transcribe_audio(entries: list[dict], downloaded: dict, progress: dict, model_size: str):
    """Transkribiert alle heruntergeladenen Audio-Dateien."""
    from faster_whisper import WhisperModel

    device, _ = detect_device()
    compute_type = "float16" if device == "cuda" else "int8"

    print(f"\n{'='*60}")
    print(f"PHASE 3: Transkription")
    print(f"  Modell: {model_size} | Geraet: {device} | Compute: {compute_type}")
    print(f"  Sprache: {LANGUAGE}")
    print(f"{'='*60}\n")

    print("  Lade Whisper-Modell (beim ersten Mal wird es heruntergeladen)...")
    model = WhisperModel(model_size, device=device, compute_type=compute_type)
    print("  Modell geladen.\n")

    unique_ids_to_transcribe = set()
    for e in entries:
        vid = e["vimeo_id"]
        if vid and vid in downloaded and downloaded[vid].get("status") == "ok":
            unique_ids_to_transcribe.add(vid)

    total = len(unique_ids_to_transcribe)
    transcribed = progress.get("transcribed", {})
    skipped = 0
    failed = 0
    new_transcriptions = 0
    total_audio_duration = 0

    start_time = time.time()

    for i, vimeo_id in enumerate(sorted(unique_ids_to_transcribe), 1):
        if vimeo_id in transcribed and transcribed[vimeo_id].get("status") == "ok":
            skipped += 1
            print(f"  [{i}/{total}] {vimeo_id} -- bereits transkribiert, uebersprungen")
            continue

        audio_path = downloaded[vimeo_id]["path"]
        if not os.path.exists(audio_path):
            transcribed[vimeo_id] = {"status": "error", "error": "Audio-Datei nicht gefunden"}
            failed += 1
            continue

        # Uebungsname fuer Anzeige finden
        uebung_name = vimeo_id
        for e in entries:
            if e["vimeo_id"] == vimeo_id:
                uebung_name = e["uebung"][:50]
                break

        print(f"  [{i}/{total}] Transkribiere: {uebung_name} ...", end=" ", flush=True)

        try:
            t0 = time.time()
            segments, info = model.transcribe(
                audio_path,
                language=LANGUAGE,
                beam_size=5,
                vad_filter=True,
                vad_parameters=dict(min_silence_duration_ms=500),
            )

            text_parts = []
            for segment in segments:
                text_parts.append(segment.text.strip())

            full_text = " ".join(text_parts)
            duration = time.time() - t0
            total_audio_duration += info.duration

            transcribed[vimeo_id] = {
                "status": "ok",
                "text": full_text,
                "audio_duration_s": round(info.duration, 1),
                "processing_time_s": round(duration, 1),
            }
            new_transcriptions += 1

            speed = info.duration / duration if duration > 0 else 0
            print(f"OK ({format_duration(info.duration)} Audio in {format_duration(duration)}, {speed:.1f}x)")

        except Exception as ex:
            transcribed[vimeo_id] = {"status": "error", "error": str(ex)[:200]}
            failed += 1
            print(f"FEHLER: {str(ex)[:80]}")

        progress["transcribed"] = transcribed
        save_progress(PROGRESS_FILE, progress)

    elapsed = time.time() - start_time
    print(f"\n  Ergebnis: {new_transcriptions} neu | {skipped} uebersprungen | {failed} fehlgeschlagen")
    print(f"  Gesamt-Audiozeit: {format_duration(total_audio_duration)}")
    print(f"  Verarbeitungszeit: {format_duration(elapsed)}")

    return transcribed


# ============================================================================
# PHASE 4: ERGEBNIS-EXPORT
# ============================================================================


def export_results(entries: list[dict], transcribed: dict, downloaded: dict):
    """Schreibt die Ergebnisse als Excel und CSV."""
    print(f"\n{'='*60}")
    print(f"PHASE 4: Ergebnis-Export")
    print(f"{'='*60}\n")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transkripte"

    headers = [
        "Order", "Modul", "Bereich", "Kategorie", "Uebung",
        "Link_Typ", "URL", "Vimeo_ID", "Audio_Dauer_s", "Transkript", "Status"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    csv_lines = [";".join(headers)]

    ok_count = 0
    fail_count = 0

    for row_idx, entry in enumerate(entries, 2):
        vid = entry["vimeo_id"]

        # Status bestimmen
        if not vid:
            status = "keine_vimeo_id"
            text = ""
            audio_dur = ""
        elif vid in transcribed and transcribed[vid].get("status") == "ok":
            status = "ok"
            text = transcribed[vid]["text"]
            audio_dur = transcribed[vid].get("audio_duration_s", "")
            ok_count += 1
        elif vid in downloaded and downloaded[vid].get("status") == "error":
            status = "download_fehlgeschlagen"
            text = f"[Fehler: {downloaded[vid].get('error', '?')}]"
            audio_dur = ""
            fail_count += 1
        elif vid in transcribed and transcribed[vid].get("status") == "error":
            status = "transkription_fehlgeschlagen"
            text = f"[Fehler: {transcribed[vid].get('error', '?')}]"
            audio_dur = ""
            fail_count += 1
        else:
            status = "nicht_verarbeitet"
            text = ""
            audio_dur = ""
            fail_count += 1

        values = [
            entry.get("order", ""),
            entry.get("modul", ""),
            entry.get("bereich", ""),
            entry.get("kategorie", ""),
            entry.get("uebung", ""),
            entry.get("link_typ", ""),
            entry.get("url", ""),
            vid or "",
            audio_dur,
            text,
            status,
        ]

        for col, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=val)

        csv_line = ";".join(
            '"' + str(v).replace('"', '""') + '"' if ";" in str(v) or '"' in str(v) or "\n" in str(v)
            else str(v)
            for v in values
        )
        csv_lines.append(csv_line)

    # Spaltenbreiten anpassen
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 45
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 80
    ws.column_dimensions["K"].width = 22

    wb.save(OUTPUT_EXCEL)
    print(f"  Excel gespeichert: {OUTPUT_EXCEL}")

    with open(OUTPUT_CSV, "w", encoding="utf-8-sig") as f:
        f.write("\n".join(csv_lines))
    print(f"  CSV gespeichert:   {OUTPUT_CSV}")

    print(f"\n  Ergebnis: {ok_count} OK | {fail_count} fehlgeschlagen/offen")
    print(f"  Eintraege gesamt:  {len(entries)}")


# ============================================================================
# MAIN
# ============================================================================


def find_excel_file() -> str:
    """Sucht die Excel-Datei im aktuellen Verzeichnis oder im Downloads-Ordner."""
    candidates = [
        os.path.join(".", EXCEL_FILENAME),
        os.path.join(os.path.expanduser("~"), "Downloads", EXCEL_FILENAME),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path

    # Fallback: suche nach *v5*.xlsx im aktuellen Verzeichnis
    for f in os.listdir("."):
        if "v5" in f and f.endswith(".xlsx") and not f.startswith("~"):
            return f

    return ""


def main():
    parser = argparse.ArgumentParser(description="Vimeo Transkript Generator")
    parser.add_argument("--skip-download", action="store_true",
                        help="Audio-Download ueberspringen (setzt vorhandene Audio-Dateien voraus)")
    parser.add_argument("--skip-transcribe", action="store_true",
                        help="Transkription ueberspringen (nur Audio herunterladen)")
    parser.add_argument("--model", default=DEFAULT_MODEL,
                        help=f"Whisper-Modell (default: {DEFAULT_MODEL}). Optionen: tiny, small, medium, large-v3")
    parser.add_argument("--excel", default="",
                        help="Pfad zur Excel-Datei (wird automatisch gesucht wenn nicht angegeben)")
    args = parser.parse_args()

    print("=" * 60)
    print("  VIMEO TRANSKRIPT GENERATOR")
    print("=" * 60)

    # Excel finden
    excel_path = args.excel or find_excel_file()
    if not excel_path or not os.path.exists(excel_path):
        print(f"\n  FEHLER: Excel-Datei nicht gefunden!")
        print(f"  Gesucht: {EXCEL_FILENAME}")
        print(f"  Lege die Datei ins aktuelle Verzeichnis oder nutze --excel <pfad>")
        sys.exit(1)

    print(f"\n  Excel: {os.path.abspath(excel_path)}")

    # Verzeichnisse erstellen
    os.makedirs(AUDIO_DIR, exist_ok=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Fortschritt laden
    progress = load_progress(PROGRESS_FILE)

    # PHASE 1: Excel einlesen
    print(f"\n{'='*60}")
    print(f"PHASE 1: Excel einlesen")
    print(f"{'='*60}\n")

    entries = read_excel(excel_path)
    unique_ids = set(e["vimeo_id"] for e in entries if e["vimeo_id"])

    print(f"  Sheet: {SHEET_NAME}")
    print(f"  Eintraege (Links): {len(entries)}")
    print(f"  Einzigartige Vimeo-IDs: {len(unique_ids)}")
    print(f"  Davon Video_kurz: {sum(1 for e in entries if e['link_typ'] == 'kurz')}")
    print(f"  Davon Video_lang: {sum(1 for e in entries if e['link_typ'] == 'lang')}")

    # PHASE 2: Audio-Download
    downloaded = progress.get("downloaded", {})
    if not args.skip_download:
        downloaded = download_audio(entries, progress)
    else:
        print(f"\n  Audio-Download uebersprungen (--skip-download)")
        # Bestehende Audio-Dateien erkennen
        for f in os.listdir(AUDIO_DIR) if os.path.exists(AUDIO_DIR) else []:
            if f.endswith(".mp3"):
                vid = f.replace(".mp3", "")
                if vid not in downloaded:
                    downloaded[vid] = {"status": "ok", "path": os.path.join(AUDIO_DIR, f)}
        print(f"  Gefundene Audio-Dateien: {sum(1 for v in downloaded.values() if v.get('status') == 'ok')}")

    # PHASE 3: Transkription
    transcribed = progress.get("transcribed", {})
    if not args.skip_transcribe:
        transcribed = transcribe_audio(entries, downloaded, progress, args.model)
    else:
        print(f"\n  Transkription uebersprungen (--skip-transcribe)")

    # PHASE 4: Export
    export_results(entries, transcribed, downloaded)

    print(f"\n{'='*60}")
    print(f"  FERTIG!")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
